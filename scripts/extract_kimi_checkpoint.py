"""Extract the inspected first consolidated checkpoint; never save a source workbook.

Workbook-specific mappings intentionally live here, not in Skills.
"""
import argparse
from collections import Counter, defaultdict
import json
from pathlib import Path
import re

import openpyxl

from validate_research import ROOT
from research_schema import apply_entry_verification_policy

MASTER = 'WHRG_2026_Master.xlsx'
FIELD = 'WHRG_2026_Field_Photo_Transcriptions_SearchSeeds.md'
PHASE1 = 'competition-entry-phase1.xlsx'
C005 = 'competition-entry-phase2-c005.xlsx'
BATCH = 'competition-entry-phase2-batch01.xlsx'
DELIVERIES = [MASTER, FIELD, PHASE1, C005, BATCH]
UNKNOWN = {None, '', 'N/A', 'None', 'Unknown', '[Unknown]', 'unknown'}
REF_PATTERN = re.compile(r'(?<![A-Za-z0-9-])(?:EV-[A-Za-z0-9-]+|FP-\d{3}|UF-\d{3})(?![A-Za-z0-9-])')
RANGE_PATTERN = re.compile(r'(EV-)(\d{3})~EV-(\d{3})')

def dump(path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + '\n', encoding='utf-8', newline='\n')

def metadata(value):
    return 'unknown' if value in UNKNOWN else str(value)

def split_evidence(text):
    text = str(text or '')
    def expand(match):
        return ','.join(match[1] + str(i).zfill(3) for i in range(int(match[2]), int(match[3]) + 1))
    return list(dict.fromkeys(REF_PATTERN.findall(RANGE_PATTERN.sub(expand, text))))

def read_books(paths):
    books, schema = {}, {}
    for filename in DELIVERIES:
        if not filename.endswith('.xlsx'):
            continue
        cached = openpyxl.load_workbook(ROOT / paths[filename], read_only=True, data_only=True)
        raw = openpyxl.load_workbook(ROOT / paths[filename], read_only=True, data_only=False)
        books[filename], schema[filename] = {}, {}
        for ws in cached:
            table = {}
            for number, cells in enumerate(ws, 1):
                values = [c.value for c in cells]
                if not any(v is not None for v in values):
                    continue
                if any(c.data_type == 'f' for c in raw[ws.title][number]):
                    raise ValueError(f'Formula found; explicit cached-value review required: {filename}/{ws.title}/{number}')
                table[number] = values
            books[filename][ws.title] = table
            schema[filename][ws.title] = {
                'nonempty_rows': list(table),
                'width': ws.max_column,
                'formulas': 0,
            }
        cached.close()
        raw.close()
    return books, schema

def extract(output):
    manifest = json.loads((ROOT / 'research/imported/kimi/manifest.json').read_text(encoding='utf-8'))
    paths = {}
    for name in DELIVERIES:
        matches = [m['path'] for m in manifest if m['original_name'] == name]
        if len(matches) != 1:
            raise ValueError(f'Expected one screened snapshot for {name}, found {len(matches)}')
        paths[name] = matches[0]
    books, schema = read_books(paths)
    entities, supplemental = [], []
    index, used_rows = {}, set()
    id_inventory = []
    def source(filename, sheet, number, identity, id_column=0, headers=None):
        values = books[filename][sheet][number]
        if identity is not None:
            if values[id_column] != identity or not isinstance(identity, str):
                raise ValueError(f'ID is not exact text: {filename}/{sheet}/{number}')
            # All identifiers were inspected as text with General number format.
            id_inventory.append({'path': paths[filename], 'sheet': sheet, 'row': number,
                                 'column': id_column + 1, 'id': identity})
        used_rows.add((filename, sheet, number))
        width = len(values)
        end = openpyxl.utils.get_column_letter(width)
        return {'path': paths[filename], 'locator': f"'{sheet}'!A{number}:{end}{number}",
                'source_id': identity, 'sheet': sheet, 'row': number,
                'id_column': id_column + 1 if identity is not None else None,
                'headers': headers, 'values': values}
    def row(filename, sheet, number, identity, header=3, id_column=0):
        result = source(filename, sheet, number, identity, id_column, books[filename][sheet].get(header))
        result['header_row'] = header if result['headers'] is not None else None
        return result
    def provenance(s):
        return {k:s[k] for k in ('path','locator','source_id')}
    def refs(item, text, reason=None):
        for identity in split_evidence(text):
            ref = {'entity_type':'Evidence','id':identity}
            if reason or ('Evidence',identity) not in index:
                issue = {'entity_type':'Evidence','id':identity,
                         'reason':reason or 'Evidence ID absent from delivered Evidence Map and field notes.'}
                if issue not in item['unresolved_references']:
                    item['unresolved_references'].append(issue)
            elif ref not in item['evidence_refs']:
                item['evidence_refs'].append(ref)
    def base(identity, label, s):
        return {'id':identity, 'label':str(label or identity), 'status':'unverified',
                'provenance':[provenance(s)], 'evidence_refs':[], 'relationships':[],
                'unresolved_references':[], 'source_rows':[s]}
    def entity(kind, identity, label, s, payload_name, payload):
        key=(kind,identity)
        if key in index:
            item=index[key]
            item['provenance'].append(provenance(s))
            item['source_rows'].append(s)
            if payload_name in item:
                raise ValueError(f'Duplicate entity facet: {key}/{payload_name}')
            item[payload_name]=payload
            return item
        item=base(identity,label,s)
        item.update(entity_type=kind)
        item[payload_name]=payload
        entities.append(item)
        index[key]=item
        return item
    def support(kind, identity, label, s, payload):
        item=base(identity,label,s)
        item['record_type']=kind
        item[kind]=payload
        supplemental.append(item)
        return item
    def fields(values,names):
        return dict(zip(names,values))
    def relation(item,name,kind,identity):
        if identity and (kind,identity) in index:
            item['relationships'].append({'relation':name,'target':{'entity_type':kind,'id':identity}})
        elif identity and identity not in UNKNOWN:
            item['unresolved_references'].append({'entity_type':kind,'id':identity,
                                                 'reason':'Typed target absent or identity unresolved.'})
    # Evidence is loaded before all dependent entities.
    for n,v in books[MASTER]['Evidence Map'].items():
        if n<4: continue
        item=entity('Evidence',v[0],v[3] or v[0],row(MASTER,'Evidence Map',n,v[0]),
                    'evidence',fields(v[1:],['topic','publisher','title','date','source_type','url',
                                           'original_text','japanese_summary','scope','confidence','notes']))
        item['source']={k:metadata(v[i]) for k,i in [('title',3),('publisher',2),('date',4),('url',6)]}
        item['source'].update(accessed_at='unknown',version='unknown',
            verification_note='Imported Kimi research; underlying source not independently checked during import.')
        item['evidence_kind']='field_evidence' if v[5]=='Field' else ('search_audit' if v[5]=='Research-Self' else 'source_record')
    lines=(ROOT/paths[FIELD]).read_text(encoding='utf-8').splitlines()
    headings=[(i,re.match(r'^## ((?:FP|UF)-\d{3})\s+.*$',line)) for i,line in enumerate(lines)]
    headings=[(i,m) for i,m in headings if m]
    for i,m in headings:
        stop=next((j for j in range(i+1,len(lines)) if lines[j].startswith('#')),len(lines))
        identity=m[1]
        s={'path':paths[FIELD], 'locator':f'lines {i+1}-{stop}', 'source_id':identity,
           'line_start':i+1,'line_end':stop,'text':'\n'.join(lines[i:stop])}
        item=entity('Evidence',identity,lines[i][3:],s,'field_evidence',
                    {'text':s['text'],'classification':'Field Evidence / search lead',
                     'underlying_photos_delivered':False})
        item['evidence_kind']='field_evidence'
        item['source']={'title':lines[i][3:],'publisher':'unknown','date':'unknown',
                        'url':'unknown','accessed_at':'unknown','version':'unknown',
                        'verification_note':'Field transcription or user observation; search lead only. Source photos not delivered; no verified web claim.'}
    # Reconnect the obsolete field-file locator without changing its supplied value.
    index[('Evidence','EV-ESM-042')]['evidence_refs']=[{'entity_type':'Evidence','id':'FP-001'}]
    # Master entities, preserving separate sheet facets rather than flattening.
    for n,v in books[MASTER]['Competition Master'].items():
        if n<4: continue
        item=entity('Competition',v[0],v[1],row(MASTER,'Competition Master',n,v[0]),'competition',
                    fields(v[1:],['official_name','category','major_category','group','competition_type',
                                  'dates','venue','heats_or_matches','medal_event','evidence_text','notes']))
        refs(item,v[10])
        if not item['evidence_refs']:
            item['evidence_gap']='Supplied evidence IDs are unresolved; no substitute source invented.'
    for n,v in books[MASTER]['Team Map'].items():
        if n<4: continue
        item=entity('Team',v[0],v[1],row(MASTER,'Team Map',n,v[0]),'team',
                    fields(v[1:],['official_name','country_region','organization_name','organization_type',
                    'competition_category','robot_platform_name','robot_manufacturer','robot_ownership',
                    'hardware_modification','software_developer','team_size','member_composition',
                    'development_start_date','development_duration','previous_robotics_experience',
                    'previous_competition_experience','training_location','hardware_support',
                    'technical_training','organizer_support','company_support','university_support',
                    'financial_support','participation_reason','expected_benefit','competition_result',
                    'urls','evidence_text','confidence','unknown_fields','last_updated','notes']))
        refs(item,v[28])
    for n,v in books[MASTER]['Recruitment Map'].items():
        if not 5<=n<=13: continue
        payload=fields(v[1:8],['team_name','formation_type','initiator','recruitment_channel',
                               'robot_source','support_received','evidence_text'])
        if v[0] in ['DOM-004','DOM-005']:
            item=support('recruitment_cohort',v[0],v[1],row(MASTER,'Recruitment Map',n,v[0],4),payload)
            item['mapping_gap']='A manufacturer customer cohort is not a single identifiable Team.'
        else:
            item=entity('Team',v[0],v[1],row(MASTER,'Recruitment Map',n,v[0],4),'recruitment',payload)
        refs(item,v[7])
    organizations={'BIC','AGIBOT','UNITREE','GALAXEA'}
    for n,v in books[MASTER]['Open-Knowledge Matrix'].items():
        if n<4: continue
        payload=fields(v[1:],['name','source_code','robot_software','model_weights','dataset',
                    'simulation','technical_report','paper','presentation','technical_blog',
                    'tutorial','video','failure_analysis','retrospective','urls','evidence_text',
                    'confidence','notes','whrg_relation','relation_evidence','publication_timing','timing_evidence'])
        if v[0]=='PKU-EPIC':
            item=support('identity_mapping',v[0],v[0],row(MASTER,'Open-Knowledge Matrix',n,v[0]),
                         {'candidate_types':['Organization','Resource'],'open_knowledge':payload,
                          'reason':'Source says Organization/Project; cannot choose one identity safely.'})
        else:
            kind='Organization' if v[0] in organizations else 'Team'
            item=entity(kind,v[0],v[0] if kind=='Organization' else v[1],
                        row(MASTER,'Open-Knowledge Matrix',n,v[0]),'open_knowledge',payload)
        refs(item,v[16])
        refs(item,v[20])
    for n,v in books[MASTER]['Rule Version Map'].items():
        if n<4: continue
        item=entity('Rule Version',v[0],v[3],row(MASTER,'Rule Version Map',n,v[0]),
                    'rule_version',fields(v[1:],['competition_ids_text','published_date','version_batch',
                    'source_url','publication_status','major_changes','previous_version','final_version','evidence_text']))
        refs(item,v[9])
        if not item['evidence_refs']: item['evidence_gap']='Rule document URL is supplied but its Evidence ID has no delivered Evidence record.'
    for item in entities:
        if item['entity_type']=='Rule Version':
            p=item['rule_version']
            relation(item,'previous_version','Rule Version',p['previous_version'])
            ids=re.fullmatch(r'C-(\d{3})\D+C-(\d{3})',p['competition_ids_text'])
            targets=[f'C-{i:03d}' for i in range(int(ids[1]),int(ids[2])+1)] if ids else [p['competition_ids_text']]
            for target in targets: relation(item,'applies_to_as_reported','Competition',target)
    # Entry identity and result data come from the Master.
    for n,v in books[MASTER]['Competition Entry Map'].items():
        if n<5: continue
        payload=fields(v[1:],['competition_id','team_id','organization_id','country','team_name',
            'organization_name','organization_type','robot_id','robot_manufacturer','robot_model',
            'robot_ownership','control_mode','result','ranking','evidence_text'])
        item=entity('Competition Entry',v[0],v[5],row(MASTER,'Competition Entry Map',n,v[0]),'entry',payload)
        refs(item,v[15])
        relation(item,'entered_in','Competition',v[1])
        for kind,i in [('Team',2),('Organization',3),('Robot Platform',8)]:
            relation(item,{'Team':'represented_by','Organization':'organization','Robot Platform':'uses_platform'}[kind],kind,v[i])
        item['identity_gaps']=['Team, Organization, and Robot Platform IDs are absent. Names are retained without inferred links.']
    # Historical entry rows carry omitted state and supplementary fields, never replacements.
    for filename,sheet in [(PHASE1,'Competition Entry Map'),(C005,'C-005 Entries'),(BATCH,'Batch01 Entries')]:
        h=books[filename][sheet][1]
        for n,v in books[filename][sheet].items():
            if n==1: continue
            d=dict(zip(h,v))
            item=index.get(('Competition Entry',v[0]))
            if item is None:
                raise ValueError(f'Historical entry absent from Master needs separate review: {v[0]}')
            s=row(filename,sheet,n,v[0],1)
            item['provenance'].append(provenance(s))
            item['source_rows'].append(s)
            reported=d['Verification Status']
            if reported not in ['Verified','Research Lead']:
                raise ValueError(f'Unknown entry status {reported}')
            item['verification']={'reported_status':reported,'authority':'Kimi',
                'independent_status':'not_checked','origin':provenance(s),
                'note':'Master has no verification column; recovered from the matching historical ID. Not independent web verification.'}
            item['entry_history']=d
    # Historical evidence with the same ID is a provenance facet, not another source.
    for filename in [PHASE1,BATCH]:
        for n,v in books[filename]['Evidence'].items():
            if n==1: continue
            item=index.get(('Evidence',v[0]))
            if item is None: raise ValueError(f'Unreviewed historical Evidence: {v[0]}')
            s=row(filename,'Evidence',n,v[0],1)
            item['provenance'].append(provenance(s))
            item['source_rows'].append(s)
    # Questions contain two incompatible layouts inside a single Master sheet.
    for n,v in books[MASTER]['Unknown Questions'].items():
        if n<4: continue
        if v[0].startswith('Q-'):
            payload=fields(v[1:],['question','category','priority','assigned_to','reported_status','notes','last_updated'])
        elif v[0].startswith('UQ-'):
            payload=fields(v[1:],['topic','question','related_competitions','evidence_text','reported_status','notes','last_updated'])
        else: raise ValueError(f'Unknown question layout: {v[0]}')
        s=row(MASTER,'Unknown Questions',n,v[0])
        item=support('question',v[0],payload['question'],s,payload)
        item['layout']='master_question' if v[0].startswith('Q-') else 'appended_phase1_question'
        refs(item,payload.get('evidence_text') or payload.get('notes'))
    qindex={i['id']:i for i in supplemental if i['record_type']=='question'}
    for n,v in books[PHASE1]['Unknown Questions'].items():
        if n==1: continue
        s=row(PHASE1,'Unknown Questions',n,v[0],1)
        if v[0] in qindex:
            item=qindex[v[0]]
            item['provenance'].append(provenance(s)); item['source_rows'].append(s)
        else:
            item=support('question',v[0],v[2],s,fields(v[1:],['topic','question','related_competitions','evidence_text','reported_status']))
            item['origin_role']='historical_record_absent_from_master'
            refs(item,v[4])
    for item in supplemental:
        if item['record_type']=='question':
            for target in re.findall(r'C-\d{3}',item['question'].get('related_competitions') or ''):
                relation(item,'asks_about','Competition',target)
    # Audits are not License, Dataset, or Resource identities.
    for n,v in books[MASTER]['License Audit'].items():
        if n<4: continue
        payload=fields(v[1:],['entity_name','asset_type','asset_name','license_found','spdx',
                    'osi_approved','commercial_use','redistribution','modification','model_training',
                    'license_url','license_text_location','evidence_text','result','notes','whrg_relation',
                    'relation_evidence','publication_timing','timing_evidence'])
        kind='license_audit_summary' if v[2]=='Summary' else 'license_audit'
        item=support(kind,v[0],v[3],row(MASTER,'License Audit',n,v[0]),payload)
        if kind=='license_audit':
            item['asset_mention']={'entity_type':'Dataset' if v[2]=='Data' else 'Resource','id':None,'label':v[3]}
            item['license_mention']={'entity_type':'License','id':None,'label':v[4]}
            item['mapping_gap']='Audit ID is preserved as an audit, not reassigned to an asset or rights instrument. Stable asset/license IDs not supplied.'
            refs(item,v[13]); refs(item,v[17])
    for sheet,kind,names in [
        ('Important Rule Changes','rule_change',['competition_text','old_rule','new_rule','effective_date','technical_implication','evidence_text','reported_status']),
        ('Case Study Summary','case_study',['organization','formation_model','key_finding','evidence_text','recovery_status'])]:
        for n,v in books[MASTER][sheet].items():
            if n<4: continue
            item=support(kind,v[0],v[1],row(MASTER,sheet,n,v[0]),fields(v[1:],names))
            if kind=='case_study':
                refs(item,v[4],'Recovered case references belong to an undelivered Evidence Log or conflict with current Evidence Map topics; not auto-linked.')
            else:
                item['source_key_gaps']=[{'text':v[6],'reason':'Shorthand source keys have no explicit mapping to delivered Evidence IDs.'}]
    for n,v in books[MASTER]['Open-Data Audit'].items():
        if 4<=n<=27:
            item=support('open_data_audit',None,v[0],row(MASTER,'Open-Data Audit',n,None),
                    fields(v,['audit_item','finding','confidence','notes']))
            refs(item,v[3])
        elif 31<=n<=34:
            support('open_data_state',None,v[0],row(MASTER,'Open-Data Audit',n,None,30),
                    fields(v,['state','definition','finding','notes']))
    for n,v in books[MASTER]['Participant Census'].items():
        if n<4: continue
        item=support('census_metric',None,v[0],row(MASTER,'Participant Census',n,None),
                     fields(v,['metric','value','source_keys','confidence','notes']))
        item['source_key_gaps']=[{'text':v[2],'reason':'No explicit mapping from source shorthand to a delivered Evidence ID.'}]
    # History-only coverage and recovery assessments remain separately attributed.
    for n,v in books[PHASE1]['Entry Source Map'].items():
        if n==1: continue
        item=support('entry_source_assessment',v[0],v[1],row(PHASE1,'Entry Source Map',n,v[0],1),
                     fields(v[1:18],['competition_name','category','registration_list_found','schedule_found',
                     'grouping_found','start_list_found','results_found','ranking_found','award_list_found',
                     'official_live_data_found','best_source_type','best_source_url','source_date','evidence_text',
                     'coverage_type','recoverability','notes']))
        refs(item,v[14]); relation(item,'assesses','Competition',v[0])
        item['origin_role']='historical_assessment_not_in_master'
    for n,v in books[BATCH]['Summary'].items():
        if n==1: continue
        item=support('entry_recovery_summary',v[0],v[1],row(BATCH,'Summary',n,v[0],1),
                     fields(v[1:],['competition_name','coverage_before','coverage_after','entries_reconstructed',
                     'finished','dns_dnf_dq','recoverability','new_primary_sources','unresolved','notes']))
        relation(item,'assesses','Competition',v[0])
        item['origin_role']='historical_assessment_not_in_master'
    values=books[C005]['C-005 Summary']
    s=row(C005,'C-005 Summary',2,values[2][1],1,1)
    item=support('competition_count_summary',values[2][1],values[2][1],s,
                 {v[0]:v[1] for n,v in values.items() if n>1})
    for n,v in values.items():
        if n>2:
            extra=row(C005,'C-005 Summary',n,None,1)
            item['source_rows'].append(extra)
    relation(item,'assesses','Competition',item['id'])
    item['origin_role']='historical_assessment_not_in_master'
    for n,v in books[BATCH]['Validation'].items():
        if n==1: continue
        support('historical_validation',None,v[0],row(BATCH,'Validation',n,None,1),
                fields(v,['check','reported_result','notes']))
    # Preserve all context and placeholders, without counting them as research entities.
    for filename,sheets in books.items():
        for sheet,table in sheets.items():
            remaining=[n for n in table if (filename,sheet,n) not in used_rows]
            if remaining:
                first=remaining[0]
                item=support('workbook_context',None,f'{filename}: {sheet}',
                    row(filename,sheet,first,None,1),
                    {'role':'Instructions, headers, templates, or recovery history; source content only.'})
                for n in remaining[1:]:
                    item['source_rows'].append(row(filename,sheet,n,None,1))
    # Same competition/name/result combinations are only possible duplicates, not aliases.
    duplicates=defaultdict(list)
    for item in entities:
        if item['entity_type']=='Competition Entry':
            p=item['entry']
            duplicates[(p['competition_id'],p['team_name'],p['result'],p['ranking'])].append(item['id'])
    groups=[ids for ids in duplicates.values() if len(ids)>1]
    for ids in groups:
        for identity in ids:
            index[('Competition Entry',identity)]['possible_duplicate_ids']=[x for x in ids if x!=identity]
    # Every row remains attributable, including unresolved source keys.
    for item in entities:
        if item['entity_type']=='Competition':
            text=item['competition']['evidence_text']
            shorthand=[x.strip() for x in str(text).split(';') if x.strip() and not x.strip().startswith('EV-')]
            if shorthand: item['source_key_gaps']=[{'text':x,'reason':'No explicit Evidence ID crosswalk delivered.'} for x in shorthand]
    apply_entry_verification_policy(entities)
    canonical_status=Counter(i['verification']['canonical_status'] for i in entities if i['entity_type']=='Competition Entry')
    counts=Counter(i['entity_type'] for i in entities)
    status=Counter(i['verification']['reported_status'] for i in entities if i['entity_type']=='Competition Entry')
    questions=Counter(i['question']['reported_status'] for i in supplemental if i['record_type']=='question')
    stats={'entity_counts':dict(sorted(counts.items())),
           'supplemental_counts':dict(sorted(Counter(i['record_type'] for i in supplemental).items())),
           'reported_entry_status':dict(status),'canonical_entry_status':dict(canonical_status),
           'independently_verified_entities':0,
           'question_status':dict(questions),'possible_duplicate_groups':groups,
           'source_id_occurrences':len(id_inventory),'distinct_source_ids':len({i['id'] for i in id_inventory}),
           'master_rows':{'Competition Entry Map':sum(n>=5 for n in books[MASTER]['Competition Entry Map']),
                          'Evidence Map':sum(n>=4 for n in books[MASTER]['Evidence Map']),
                          'Unknown Questions':sum(n>=4 for n in books[MASTER]['Unknown Questions'])},
           'unresolved_reference_ids':sorted({r['id'] for i in entities+supplemental for r in i['unresolved_references']})}
    # Inspect raw ID cell types/formats again in the immutable workbook, no normalization.
    for filename in books:
        wb=openpyxl.load_workbook(ROOT/paths[filename],read_only=True,data_only=False)
        for loc in id_inventory:
            if loc['path']!=paths[filename]: continue
            cell=wb[loc['sheet']].cell(loc['row'],loc['column'])
            if cell.data_type!='s' or cell.value!=loc['id'] or cell.number_format!='General':
                raise ValueError(f'Ambiguous identifier cell: {loc}')
        wb.close()
    # Markdown IDs use exact heading capture rather than fabricated source IDs.
    for item in entities:
        if 'field_evidence' in item:
            s=item['source_rows'][0]
            id_inventory.append({'path':s['path'],'line':s['line_start'],'id':item['id']})
    stats['source_id_occurrences']=len(id_inventory)
    stats['distinct_source_ids']=len({i['id'] for i in id_inventory})
    for filename,sheets in books.items():
        for sheet,table in sheets.items():
            header_numbers = [1] if filename != MASTER else [3]
            if filename == MASTER and sheet == 'Recruitment Map': header_numbers = [4,17]
            if filename == MASTER and sheet == 'Instructions': header_numbers = []
            schema[filename][sheet]['headers'] = {str(n): table[n] for n in header_numbers if n in table}
    schema['policy']={'read_only':True,'data_only':True,'identifier_type':'text','identifier_number_format':'General',
                      'master_entry_header_row':3,'master_entry_template_row':4,
                      'master_question_layouts':{'4-31':'Q layout','32-34':'phase1 layout with last-updated column'}}
    dump(output/'candidate.json',entities)
    dump(output/'supplemental.json',supplemental)
    dump(output/'checkpoint-schema.json',schema)
    dump(output/'checkpoint-id-inventory.json',id_inventory)
    dump(output/'checkpoint-stats.json',stats)
    print(json.dumps(stats,ensure_ascii=True,indent=2))

def main():
    parser=argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--output',type=Path,default=ROOT/'research/staging')
    args=parser.parse_args()
    output=args.output.resolve()
    if not output.is_relative_to((ROOT/'research/staging').resolve()):
        raise ValueError('Extraction must target repository staging; promotion is a separate reviewed step.')
    extract(output)

if __name__=='__main__':
    main()
